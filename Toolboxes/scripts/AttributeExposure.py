import arcpy
import os
import time
import xlrd
import openpyxl

import scripts.common_lib as common_lib
from scripts.common_lib import create_msg_body, msg, trace
from scripts.settings import *

# error classes
class NotProjected(Exception):
    pass


class NoNoDataError(Exception):
    pass


class LicenseError3D(Exception):
    pass


class LicenseErrorSpatial(Exception):
    pass


class NoUnits(Exception):
    pass


class NoPolygons(Exception):
    pass


class SchemaLock(Exception):
    pass


class NotSupported(Exception):
    pass


class NoLayerFile(Exception):
    pass


class FunctionError(Exception):
    pass


#########################
# Definitions/Functions #
#########################


def processByRiskType(riskType):
    if riskType == "NOAA Sea Level Rise":
        print("NOAA")
    if riskType == "FEMA Flood":
        print("FEMA Flood")
    if riskType == "Tidal Food":
        print("Tidal Flood")
    if riskType == "Storm Surge":
        print("Storm Surge")


def riskTypeValues(riskType, inSpreadsheet):
    ''' Returns Risk Attribute Name and attribute values to process for each exposure/risk level'
    ** the attacehd riskTypeTable.xlsx is called-to ** User can add additional exposure levels
    in the spreadsheet in Max to Min order '''
    from openpyxl import load_workbook
    workbook = load_workbook(inSpreadsheet)
    worksheet = workbook.get_sheet_by_name("riskLevels")
    attrBaseName = ""
    levelsList = []
    levelsListStr = []
    for row in worksheet.iter_rows():
        if row[0].value == riskType:
            count = 0
            for item in row:
                if count == 0:
                    pass
                if count == 1:
                    attrBaseName = item.value
                if count > 1:
                    if item.value is not None:
                        levelsList.append(item.value)
                count += 1
    for item in levelsList:
        val = str(item).replace(".", "_")
        levelsListStr.append(val)
    return attrBaseName, levelsList, levelsListStr


def obtainProcessingList(inRiskGeoms3D, inriskValues):
    riskGeoms3DOrdered = []
    for riskGeom3D in inRiskGeoms3D:
        skip = False
        for riskValue in inriskValues[2]:
            if skip:
                pass
            else:
                riskGeomNo3D = str(riskGeom3D).replace("3D", "").replace("3d", "")
                if riskValue in str(riskGeomNo3D):
                    riskValPosition = riskGeomNo3D.index(str(riskValue))
                    riskValueLength = len(riskValue)
                    if riskGeom3D[riskValPosition:(riskValPosition+riskValueLength)] == riskValue:
                        riskGeoms3DOrdered.append([riskValue, riskGeom3D])
                        skip = True
    return riskGeoms3DOrdered[::-1]


def obtainGDFC(inGDB, featureType):
    arcpy.env.workspace = inGDB
    # Use the ListFeatureClasses function to return a list of
    #  shapefiles.
    featureclasses = arcpy.ListFeatureClasses(feature_type=featureType)
    # Copy shapefiles to a file geodatabase
    return featureclasses


def unitsCalc(inFeature):
    SpatialRef = arcpy.Describe(inFeature).spatialReference
    obtainunits = SpatialRef.linearUnitName
    try:
        if obtainunits == "Foot_US":
            units = "Foot"
            return units
        if obtainunits == "Foot":
            units = "Foot"
            return units
        if obtainunits == "Meter":
            units = "Meter"
            return units
    except:
        arcpy.AddWarning("Units Not Detected on {0} \n Terminating Process".format(inFeature))
        exit()


def obtainExtent(inFc):
    desc = arcpy.Describe(inFc)
    xMin = desc.extent.XMin
    xMax = desc.extent.XMax
    yMin = desc.extent.YMin
    yMax = desc.extent.YMax
    extent = "{0}, {1}, {2}, {3}".format(xMin, xMax, yMin, yMax)
    return extent


def shortestExtent(inFc):
    # Detects the shortest extent of the geometry (Length or Height)
    desc = arcpy.Describe(inFc)
    xMin = desc.extent.XMin
    xMax = desc.extent.XMax
    yMin = desc.extent.YMin
    yMax = desc.extent.YMax
    length = xMax - xMin
    height = yMax - yMin
    if length >= height:
        return height
    else:
        return length


def MpRasterTool(inFeature, outRaster):
    ''' Update to the Multipatch Raster tool to '''
    val = 1  # One Meter Distance
    if unitsCalc(inFeature) == "Foot":
        val = 3.28084  # One Meter Distance Converted to Foot
    out = arcpy.MultipatchToRaster_conversion(inFeature, outRaster, val)
    return out


##########
# Script #
##########
def attribute_feature(inGDB, inBuilding, buildingFID, inDEM, outGDB, debug):
    try:
        # Get Attributes from User
        if debug == 0:
            # script variables
            aprx = arcpy.mp.ArcGISProject("CURRENT")
            home_directory = aprx.homeFolder
            tiff_directory = home_directory + "\\Tiffs"
            tin_directory = home_directory + "\\Tins"
            scripts_directory = aprx.homeFolder + "\\Scripts"
            rule_directory = aprx.homeFolder + "\\rule_packages"
            log_directory = aprx.homeFolder + "\\Logs"
            layer_directory = home_directory + "\\layer_files"
            project_ws = aprx.defaultGeodatabase

            enableLogging = True
            DeleteIntermediateData = True
            verbose = 0
            in_memory_switch = True
        else:
            # debug
            riskType = "NOAA Sea Level Rise"  # "NOAA Sea Level Rise", "FEMA Flood", "Tidal Flood", "Storm Surge"
#            inGDB = r'C:\Users\geof7015\PycharmProjects\3DFloodImpact\NOAA_SeaLevelRise.gdb'
#            inBuilding = r'C:\Users\geof7015\PycharmProjects\3DFloodImpact\Baltimore.gdb\LOD2_Buildings_test_area1'
#            buildingFID = "BuildingFID"
#            inDEM = r'C:\Users\geof7015\PycharmProjects\3DFloodImpact\Baltimore.gdb\Sandy_Baltimore_dtm_2m_test_area1_1900'
#            outGDB = r'C:\Users\geof7015\PycharmProjects\3DFloodImpact\test.gdb'

            inGDB = r'D:\Gert\Work\Esri\Solutions\3DFloodImpact\work2.1\3DFloodImpact\NOAA_SeaLevelRise.gdb'
            inBuilding = r'D:\Gert\Work\Esri\Solutions\3DFloodImpact\work2.1\3DFloodImpact\Baltimore.gdb\LOD2_Buildings_test_area1'
            buildingFID = "BuildingFID"
            inDEM = r'D:\Gert\Work\Esri\Solutions\3DFloodImpact\work2.1\3DFloodImpact\Baltimore.gdb\Sandy_Baltimore_dtm_2m_test_area1_1900'
            outGDB = r'D:\Gert\Work\Esri\Solutions\3DFloodImpact\work2.1\3DFloodImpact\Testing.gdb'

            home_directory = r'D:\Gert\Work\Esri\Solutions\3DFloodImpact\work2.1\3DFloodImpact'
            log_directory = home_directory + "\\Logs"
            layer_directory = home_directory + "\\LayerFiles"
            project_ws = home_directory + "\\Testing.gdb"

            enableLogging = False
            DeleteIntermediateData = True
            verbose = 1
            in_memory_switch = False

        scratch_ws = common_lib.create_gdb(home_directory, "Intermediate.gdb")
        arcpy.env.workspace = scratch_ws
        arcpy.env.overwriteOutput = True

        start_time = time.clock()

        if arcpy.CheckExtension("3D") == "Available":
            arcpy.CheckOutExtension("3D")

            if arcpy.CheckExtension("Spatial") == "Available":
                arcpy.CheckOutExtension("Spatial")

                arcpy.AddMessage("Processing input source: " + common_lib.get_name_from_feature_class(inBuilding))

                processByRiskType(riskType)
                inSpreadsheet = r'riskTypeTable.xlsx'
                riskValues = riskTypeValues(riskType, inSpreadsheet)
                riskAttrBaseName = riskValues[0]
                riskBaseName = riskValues[0]
                riskGeoms3D = obtainGDFC(inGDB, "Multipatch")
                # alert user what data will not be processed
                processList = obtainProcessingList(riskGeoms3D, riskValues)
                print(processList)

                tempFP = os.path.join("in_memory", "tempFP")
                arcpy.MultiPatchFootprint_3d(inBuilding, tempFP, buildingFID)

                # Check for Meters vs Feet elevations in DEM vs Water Surface Rasters
                count = 0
                for row in processList:
                    if count < 1:
                        riskGeom3D = row[1]
                        riskGeom3DFullPath = os.path.join(inGDB, riskGeom3D)
                        # tempRiskSurfaceRaster = os.path.join("in_memory", "tempRiskSurfaceRaster")
                        tempRiskSurfaceRaster = os.path.join(outGDB, "SurfaceRaster{0}{1}".format(riskValues[0], row[0]))
                        if arcpy.Exists:
                            arcpy.Delete_management(tempRiskSurfaceRaster)
                        MpRasterTool(riskGeom3DFullPath, tempRiskSurfaceRaster)
                        # Generate Flood Depth Raster
                        ''' Retains all pixels less than or equal to the flood raster and adds Flood Elevation'''
                        #arcpy.sa.Raster(inDEM) >= arcpy.sa.Raster(tempRiskSurfaceRaster)
                        #arcpy.sa.Raster(inDEM) + arcpy.sa.Raster(tempRiskSurfaceRaster)
                        #output_raster =
                        depthRaster = os.path.join(outGDB, "depthRaster{0}{1}".format(riskValues[0], row[0]))
                        if arcpy.Exists(depthRaster):
                            arcpy.Delete_management(depthRaster)
                        depthRasterOpp = arcpy.sa.SetNull(arcpy.sa.Raster(inDEM) >= arcpy.sa.Raster(tempRiskSurfaceRaster),
                                                          arcpy.sa.Raster(tempRiskSurfaceRaster) - arcpy.sa.Raster(inDEM),
                                                          "Value = 1")
                        depthRasterOpp.save(depthRaster)
                        depthZonalStatsTable = os.path.join(outGDB, "depthZonalStatsTable{0}{1}".format(riskValues[0], row[0]))
                        if arcpy.Exists(depthZonalStatsTable):
                            arcpy.Delete_management(depthZonalStatsTable)
                        arcpy.sa.ZonalStatisticsAsTable(tempFP, buildingFID, depthRaster, depthZonalStatsTable,
                                                        "DATA", "ALL")
                        # TODO Geof7015 ensure that accurate area attribute can always be obtained
                        arcpy.JoinField_management(depthZonalStatsTable, buildingFID, tempFP, buildingFID, "Shape_Area")

                        arcpy.AddField_management(depthZonalStatsTable, "Exposure{0}{1}".format(riskValues[0], row[0]), "DOUBLE", None,
                                                  None, None, "Exposure{0}{1}".format(riskValues[0], row[0]), "NULLABLE",
                                                  "NON_REQUIRED", None)

                        fields = arcpy.ListFields(tempFP)
                        for field in fields:
                            print("{0} is a type of {1} with a length of {2}"
                                  .format(field.name, field.type, field.length))
                        #'''0.) Calculate in decreasing order. Highest to lowest elevation.'''
                        ''' ---- This ensures that only buildings that are exposed will be run in later recursions to improve performance '''
                        '''1.) Alter the fields that need to be added '''
                        '''2.) Obtain Zonal Stats for the Ground Elevations first impacted'''
                        '''3.) Alter the fields that need to be added '''
                        '''4.) Merge all tables for each recursion into a master table'''
                        '''5.) Join the tables to tempFP'''
                        '''6.) Run the calculations below on tempFP (Exposure, Volume, SLR)'''
                        updateFieldsList = []
                        with arcpy.da.UpdateCursor(tempFP, updateFieldsList) as updateRows:
                            for row in updateRows:
                                print(row)
                                # Calculate Percent Exposure
                                if None in row[1]:
                                    row[2] = 0
                                else:
                                    row[2] = (100 / row[0]) * row[1]
                                row.updateRow(row)
                        # Percent Exposure = (100/"Shape_Area")* buildingFID
                        #isFloodedRaster = os.path.join(outGDB, "isFloodedRaster{0}".format(count))
                        #output_raster.save(isFloodedRaster)
                        #arcpy.Delete_management(tempRiskSurfaceRaster)
                        #arcpy.AddField_management()
                    count += 1

                end_time = time.clock()
                msg_body = create_msg_body("attribute_feature completed successfully.", start_time, end_time)
                msg(msg_body)

                arcpy.ClearWorkspaceCache_management()

                return None

            else:
                raise LicenseErrorSpatial
        else:
            raise LicenseError3D

        arcpy.ClearWorkspaceCache_management()


    except NotProjected:
        print("Input data needs to be in a projected coordinate system. Exiting...")
        arcpy.AddError("Input data needs to be in a projected coordinate system. Exiting...")

    except NoLayerFile:
        print("Can't find Layer file. Exiting...")
        arcpy.AddError("Can't find Layer file. Exiting...")

    except LicenseError3D:
        print("3D Analyst license is unavailable")
        arcpy.AddError("3D Analyst license is unavailable")

    except LicenseErrorSpatial:
        print("Spatial Analyst license is unavailable")
        arcpy.AddError("Spatial Analyst license is unavailable")

    except NoNoDataError:
        print("Input raster does not have NODATA values")
        arcpy.AddError("Input raster does not have NODATA values")

    except NoUnits:
        print("No units detected on input data")
        arcpy.AddError("No units detected on input data")

    except NoPolygons:
        print("Input data can only be polygon features or raster datasets.")
        arcpy.AddError("Input data can only be polygon features or raster datasets.")

    except ValueError:
        print("Input no flood value is not a number.")
        arcpy.AddError("Input no flood value is not a number.")

    except arcpy.ExecuteError:
        line, filename, synerror = trace()
        msg("Error on %s" % line, ERROR)
        msg("Error in file name:  %s" % filename, ERROR)
        msg("With error message:  %s" % synerror, ERROR)
        msg("ArcPy Error Message:  %s" % arcpy.GetMessages(2), ERROR)

    except FunctionError as f_e:
        messages = f_e.args[0]
        msg("Error in function:  %s" % messages["function"], ERROR)
        msg("Error on %s" % messages["line"], ERROR)
        msg("Error in file name:  %s" % messages["filename"], ERROR)
        msg("With error message:  %s" % messages["synerror"], ERROR)
        msg("ArcPy Error Message:  %s" % messages["arc"], ERROR)

    except:
        line, filename, synerror = trace()
        msg("Error on %s" % line, ERROR)
        msg("Error in file name:  %s" % filename, ERROR)
        msg("with error message:  %s" % synerror, ERROR)

    finally:
        arcpy.CheckInExtension("3D")
        arcpy.CheckInExtension("Spatial")


# for debug only!
if __name__ == "__main__":
    attribute_feature("", "", "", "", "", 1)

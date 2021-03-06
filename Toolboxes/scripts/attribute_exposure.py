import arcpy
import os
import re
import time
import sys
import flood_impact_lib
import common_lib
from common_lib import create_msg_body, msg, trace
import settings
from settings import *

# https://www.fema.gov/media-library-data/1523562952942-4c54fdae20779bb004857f1915236e6c/Flood_Depth_and_Analysis_Grids_Guidance_Feb_2018.pdf

# error classes
class NotProjected(Exception):
    pass


class MixOfSR(Exception):
    pass


class GeopraphicSR(Exception):
    pass


class NoNoDataError(Exception):
    pass


class LicenseError3D(Exception):
    pass


class LicenseErrorSpatial(Exception):
    pass


class NoUnits(Exception):
    pass


class NoPolygons(Exception):
    pass


class SchemaLock(Exception):
    pass


class NotSupported(Exception):
    pass


class NoLayerFile(Exception):
    pass


class NoRiskTable(Exception):
    pass


class FunctionError(Exception):
    pass


class InputError(Exception):
    pass


class NoLossTable(Exception):
    pass


LOSSTABLE = "LossPotential"
ERROR = "error"
WARNING = "warning"

#########################
# Definitions/Functions #
#########################

def create_point_featureclass_from_points(fc_name, points_list, spatial_ref):
    try:
        # create the output fc
        geometry_type = "POINT"
        template = ""
        has_m = "DISABLED"
        has_z = "DISABLED"

        ws_path, stats_out_name = os.path.split(fc_name)
        arcpy.CreateFeatureclass_management(ws_path, stats_out_name, geometry_type, template,
                                            has_m, has_z, spatial_ref)

        f_cursor = arcpy.da.InsertCursor(fc_name, ["SHAPE@XY"])

        for pt in points_list:
            f_cursor.insertRow(pt)

        del f_cursor

    except arcpy.ExecuteWarning:
        print((arcpy.GetMessages(1)))
        arcpy.AddWarning(arcpy.GetMessages(1))

    except arcpy.ExecuteError:
        print((arcpy.GetMessages(2)))
        arcpy.AddError(arcpy.GetMessages(2))

    # Return any other type of error
    except:
        # By default any other errors will be caught here
        e = sys.exc_info()[1]
        print((e.args[0]))
        arcpy.AddError(e.args[0])


def checkSameSpatialReference(input_list, featureclass_list):
    try:
        we_fail = 0

        base_sr = arcpy.Describe(input_list[0]).spatialReference
        base_linear_unit = base_sr.linearUnitName
        base_Zunit = common_lib.get_z_unit(input_list[0], 0)

        one_list = input_list + featureclass_list
        for f in one_list:
            if arcpy.Exists(f):
                sr = arcpy.Describe(f).spatialReference

                if sr.name != base_sr.name:
                    arcpy.AddMessage(base_sr.name)
                    arcpy.AddMessage(common_lib.get_name_from_feature_class(f) + " has different spatial reference " + sr.name + " than " + common_lib.get_name_from_feature_class(input_list[0]))
                    we_fail = 1
                    break
                else:
                    if sr.linearUnitName != base_linear_unit:
                        arcpy.AddMessage(common_lib.get_name_from_feature_class(f) + " has different linear units " + sr.linearUnitName
                                         + " than " + common_lib.get_name_from_feature_class(input_list[0]))
                        we_fail = 1
                        break
                    else:
                        if common_lib.get_z_unit(f, 0) != base_Zunit:
                            we_fail = 1
                            arcpy.AddMessage(common_lib.get_name_from_feature_class(f) + " has different spatial reference or units than " + common_lib.get_name_from_feature_class(input_list[0]))
                            break
        return we_fail

    except arcpy.ExecuteWarning:
        print((arcpy.GetMessages(1)))
        arcpy.AddWarning(arcpy.GetMessages(1))

    except arcpy.ExecuteError:
        print((arcpy.GetMessages(2)))
        arcpy.AddError(arcpy.GetMessages(2))

    # Return any other type of error
    except:
        # By default any other errors will be caught here
        e = sys.exc_info()[1]
        print((e.args[0]))
        arcpy.AddError(e.args[0])


def riskTypeValues(riskType, inSpreadsheet):
    ''' Returns Risk Attribute Name and attribute values to process for each exposure/risk level'
    ** the attached riskTypeTable.xlsx is called-to ** User can add additional exposure levels
    in the spreadsheet in Max to Min order '''
    from openpyxl import load_workbook
    workbook = load_workbook(inSpreadsheet)
    worksheet = workbook["riskLevels"]
    attrBaseName = ""
    levelsList = []
    levelsListStr = []
    isPercentFlood = False
    for row in worksheet.iter_rows():
        if row[0].value == riskType:
            count = 0
            for item in row:
                if count == 0:
                    pass
                if count == 1:
                    attrBaseName = item.value
                if count > 1:
                    if item.value is not None:
                        levelsList.append(item.value)
                count += 1
    for item in levelsList:
        if 0 < item < 1:                        # if any item < 1, we assume percentage flood
            isPercentFlood = True
        val = str(item).replace(".", "_")
        levelsListStr.append(val)

    def orderIsDescending(list):
        for i in range(len(list) - 1):
            if list[i] < list[i + 1]:
                return False
            return True

    if isPercentFlood and orderIsDescending(levelsList):
        levelsList = levelsList[::-1]
        levelsListStr = levelsListStr[::-1]
    elif not orderIsDescending(levelsList):
        levelsList = levelsList[::-1]
        levelsListStr = levelsListStr[::-1]

    return (attrBaseName, levelsList, levelsListStr), isPercentFlood


def compareDepthSurfaceLists(surfaceRasterProcessList, depthRasterProcessList):
    surfaceRasterListLength = len(surfaceRasterProcessList)
    depthRasterListLength = len(depthRasterProcessList)
    if surfaceRasterListLength != depthRasterListLength:
        arcpy.AddWarning("detected {0} Surface Rasters and {0} Depth Rasters"
                         .format(surfaceRasterListLength, depthRasterListLength))
        arcpy.AddError("Terminating Process. Please check and ensure 1 depth and 1 surface raster exist for each "
                         "water level")
        raise InputError
    else:
        surfaceRasterVals = []
        depthRasterVals = []
        for row in surfaceRasterProcessList:
            if row[0] not in surfaceRasterVals:
                surfaceRasterVals.append(row[0])
        for row in depthRasterProcessList:
            if row[0] not in depthRasterVals:
                depthRasterVals.append(row[0])
        compNumbers = [i for i, j in zip(surfaceRasterVals, depthRasterVals) if i == j]
        if len(compNumbers) != surfaceRasterListLength:
            arcpy.AddError("Check Raster Depth and Surface Rasters for inconsistencies. \n"
                             "One of more rasters do not exist for a specified water elevation or the rasters don't have the correct risk values for the chosen riskType.")
            raise InputError
        else:
            return True


def obtainGDFC(inGDB, featureType):
    arcpy.env.workspace = inGDB
    # Use the ListFeatureClasses function to return a list of Feature Classes of given type
    featureclasses = arcpy.ListFeatureClasses(feature_type=featureType)

    sorted_list = flood_impact_lib.fc_numeric_sorter(featureclasses)
    featureclasses = sorted_list

#    featureclasses.sort(reverse=True)
    # Copy Feature to a file geodatabase
    return featureclasses


def unitsCalc(inFeature):
    SpatialRef = arcpy.Describe(inFeature).spatialReference
    obtainunits = SpatialRef.linearUnitName
    try:
        if obtainunits == "Foot_US":
            units = "Foot"
            return units
        if obtainunits == "Foot":
            units = "Foot"
            return units
        if obtainunits == "Meter":
            units = "Meter"
            return units
        if obtainunits not in ["Foot_US", "Foot", "Meter"]:
            arcpy.AddError("Units Not Detected on {0} \n Terminating Process".format(inFeature))
            exit()
    except:
        arcpy.AddError("Units Not Detected on {0} \n Terminating Process".format(inFeature))
        exit()


def obtainExtent(inFc):
    desc = arcpy.Describe(inFc)
    xMin = desc.extent.XMin
    xMax = desc.extent.XMax
    yMin = desc.extent.YMin
    yMax = desc.extent.YMax
    extent = "{0}, {1}, {2}, {3}".format(xMin, xMax, yMin, yMax)
    return extent, xMin, xMax, yMin, yMax


def shortestExtent(inFc):
    # Detects the shortest extent of the geometry (Length or Height)
    desc = arcpy.Describe(inFc)
    xMin = desc.extent.XMin
    xMax = desc.extent.XMax
    yMin = desc.extent.YMin
    yMax = desc.extent.YMax
    length = xMax - xMin
    height = yMax - yMin
    if length >= height:
        return height
    else:
        return length


def MpRasterTool(inFeature, outRaster):
    ''' Update to the Multipatch Raster tool to '''
    val = 1  # One Meter Distance
    if unitsCalc(inFeature) == "Foot":
        val = 3.28084  # One Meter Distance Converted to Foot
    out = arcpy.MultipatchToRaster_conversion(inFeature, outRaster, val)
    return out


def pixelArea(inRaster):
    desc = arcpy.Describe(inRaster)
    cellX = round(float(str(desc.meanCellHeight).replace("e-", "")), 4)
    cellY = round(float(str(desc.meanCellWidth).replace("e-", "")), 4)
    pixelArea = round(cellX * cellY, 4)
    return pixelArea


def isRaster(inRaster):
    desc = arcpy.Describe(inRaster).dataType
    if desc in ['MosaicDataset', 'RasterDataset']:
        return True
    else:
        return False


def createTempFP(inFeature, bufferDistance, featureFID, tempFP):
    fcDesc = arcpy.Describe(inFeature)
    if fcDesc.shapeType == "MultiPatch":
        if bufferDistance > 0:
            mpFpTemp = os.path.join("in_memory", "mpFpTemp")

            arcpy.AddMessage("Creating footprints for: " + common_lib.get_name_from_feature_class(inFeature) + ".")

            arcpy.MultiPatchFootprint_3d(inFeature, mpFpTemp, featureFID)
            arcpy.RepairGeometry_management(mpFpTemp)

            # this adds the Shape_Area field
            arcpy.AddMessage("Buffering footprints of: " + common_lib.get_name_from_feature_class(inFeature) + ".")
            arcpy.Buffer_analysis(mpFpTemp, tempFP,
                                      "{0} {1}".format(bufferDistance, unitsCalc(mpFpTemp)), "FULL", "ROUND",
                                      "NONE", None, "PLANAR")

            arcpy.Delete_management(mpFpTemp)
        else:
            arcpy.AddMessage("Creating footprints for: " + common_lib.get_name_from_feature_class(inFeature) + ".")
            arcpy.MultiPatchFootprint_3d(inFeature, tempFP, featureFID)
            arcpy.RepairGeometry_management(tempFP)
    if fcDesc.shapeType == "Polygon":
        if bufferDistance > 0:
            arcpy.AddMessage("Buffering: " + common_lib.get_name_from_feature_class(inFeature) + ".")
            arcpy.Buffer_analysis(inFeature, tempFP,
                                  "{0} {1}".format(bufferDistance, unitsCalc(inFeature)), "FULL", "ROUND",
                                  "NONE", None, "PLANAR")
        else:
            arcpy.CopyFeatures_management(inFeature, tempFP)
    if fcDesc.shapeType == "Point":
        if bufferDistance <= 0:
            if unitsCalc(inFeature) == "Meter":
                bufferDistance = 0.3048
            else:
                bufferDistance = 1

        arcpy.AddMessage("Buffering: " + common_lib.get_name_from_feature_class(inFeature) + ".")
        arcpy.Buffer_analysis(inFeature, tempFP, "{0} {1}".format(bufferDistance, unitsCalc(inFeature)),
                              "FULL", "ROUND", "NONE", None, "PLANAR")
    if fcDesc.shapeType in ["Polyline", "Line"]:
        if bufferDistance <= 0:
            if unitsCalc(inFeature) == "Meter":
                bufferDistance = 0.3048
            else:
                bufferDistance = 1

        arcpy.AddMessage("Buffering: " + common_lib.get_name_from_feature_class(inFeature) + ".")
        arcpy.Buffer_analysis(inFeature, tempFP, "{0} {1}".format(bufferDistance, unitsCalc(inFeature)),
                              "FULL", "ROUND", "NONE", None, "PLANAR")


##########
# Script #
##########


def attribute_feature(riskType,
                        inWaterSurfaceType,
                        inSurfaceGDB,
                        inDepthGDB,
                        inFeature,
                        featureFID,
                        esri_featureID,
                        bufferDistance,
                        tolerance,
                        inDEM,
                        lossTable,
                        outTable,
                        areaField,
                        minField,
                        maxField,
                        rangeField,
                        meanField,
                        stdField,
                        volumeField,
                        shapeAreaField,
                        exposureField,
                        WaterSurfaceElevationLevel,
                        GroundMinField,
                        GroundMaxField,
                        GroundRangeField,
                        GroundMeanField,
                        GroundSTDField,
                        lossField,
                        debug,
                        lc_use_in_memory):
    try:
        # Get Attributes from User
        if debug == 0:
            # script variables

            aprx = arcpy.mp.ArcGISProject("CURRENT")
            home_directory = aprx.homeFolder
            layer_directory = home_directory + "\\layer_files"
            table_directory = home_directory + "\\tables"
            project_ws = aprx.defaultGeodatabase

            enableLogging = True
            DeleteIntermediateData = False
        else:
            # debug
            home_directory = r'D:\Gert\Work\Esri\Solutions\3DFloodImpact\work2.4\FloodImpactAnalysis'
            log_directory = home_directory + "\\Logs"
            layer_directory = home_directory + "\\layer_files"
            table_directory = home_directory + "\\tables"
            project_ws = home_directory + "\\Testing.gdb"
            DeleteIntermediateData = False
        scratch_ws = common_lib.create_gdb(home_directory, "Intermediate.gdb")
        arcpy.env.workspace = scratch_ws
        arcpy.env.overwriteOutput = True
        start_time = time.clock()
        success = False

        depthField = "Depth"
        potentialLossField = "PotentialLoss"
        sizeField = "Size"
        spaceUseField = "SpaceUse"
        loss_potential_field_list = [depthField, potentialLossField, sizeField, spaceUseField]

        arcpy.SetProgressor("default", "Pre-Calculating Data for Exposure Analysis")
        if arcpy.CheckExtension("3D") == "Available":
            arcpy.CheckOutExtension("3D")
            if arcpy.CheckExtension("Spatial") == "Available":
                arcpy.CheckOutExtension("Spatial")

                # check if input feature has a projected coordinate system (required!)
                cs_name, cs_vcs_name, projected = common_lib.get_cs_info(inFeature, 0)

                if projected:
                    # make a copy to grab the selection
                    input_source_copy = os.path.join(scratch_ws, common_lib.get_name_from_feature_class(inFeature) + "_copy")
                    if arcpy.Exists(input_source_copy):
                        arcpy.Delete_management(input_source_copy)

                    arcpy.AddMessage("Making copy of " + common_lib.get_name_from_feature_class(inFeature) + " to " + input_source_copy)

                    arcpy.CopyFeatures_management(inFeature, input_source_copy)

                    inFeature = input_source_copy

                    # create string field for featureFID. otherwise resulting raster won't be integer
                    common_lib.delete_add_field(inFeature, esri_featureID, "TEXT")

                    field_name = common_lib.get_field_name_for_fieldstring(inFeature, featureFID)

                    arcpy.CalculateField_management(inFeature, esri_featureID, "!" + field_name + "!", "PYTHON_9.3")

                    # make the featureFID the text version of featureFID
                    featureFID = esri_featureID

                    # check if all water and depth levels have the same spatial reference as the input feature class
                    list_fcs = []
                    list_surfaces_rasters = []
                    list_depth_rasters = []

                    if inSurfaceGDB:
                        list_fcs = common_lib.list_fcs_in_gdb(inSurfaceGDB, 0)
                        list_surfaces_rasters = common_lib.list_rasters_in_gdb(inSurfaceGDB, 0)
                    if inDepthGDB:
                        list_depth_rasters = common_lib.list_rasters_in_gdb(inDepthGDB, 0)

                    if checkSameSpatialReference([inFeature], list_fcs + list_surfaces_rasters + list_depth_rasters) == 1:
                        raise MixOfSR

                    # check if risk table exists...
                    inSpreadsheet = os.path.join(table_directory, 'riskTypeTable.xlsx')

                    if arcpy.Exists(inSpreadsheet):
                        arcpy.AddMessage("Reading risk table: " + inSpreadsheet + ".")
                        riskValues, isPercentFlood = riskTypeValues(riskType, inSpreadsheet)
                        #print(riskValues)

                        if len(riskValues[1]) > 0:
                            # check if loss potential table exists...
                            if len(lossTable) > 0:
                                if arcpy.Exists(lossTable):
                                    if "$" in lossTable:
                                        # parse loss tablespace use table info
                                        code, loss_gdb_table = common_lib.import_table_with_required_fields(lossTable, project_ws, LOSSTABLE, loss_potential_field_list, debug)

                                        if code == 0:
                                            pass
                                        else:
                                            msg_body = create_msg_body("Failed to import " + lossTable + "!", 0, 0)
                                            msg(msg_body, WARNING)
                                            lossTable = None
                                    else:
                                        msg_body = create_msg_body("Please browse to sheet ($) view within: " + lossTable + "!", 0, 0)
                                        msg(msg_body, WARNING)
                                        lossTable = None
                                else:
                                    msg_body = create_msg_body("Can't find: " + lossTable + "!", 0, 0)
                                    msg(msg_body, WARNING)
                                    lossTable = None

                            # check if surfaceData exists...
                            surfaceDataExists = False
                            if arcpy.Exists(inSurfaceGDB):

                                # Obtain List of Available Surface Features/Rasters in GDB
                                # check if there are multipatches first
                                riskGeoms3D = obtainGDFC(inSurfaceGDB, "Multipatch")

                                if len(riskGeoms3D) > 0:
                                    arcpy.AddMessage("Detected Input Surface Raster as Multipatch. **Multipatch Features "
                                                     "will take substantially longer than rasters to process. For processing "
                                                     "large areas we recommend using Raster Data.")

                                    # get processing list, alert user if data will not be processed: risk value needs to be present in naming.
                                    surfaceRasterProcessList, error = flood_impact_lib.obtain_processing_list(riskGeoms3D, riskValues,
                                                                                           inSurfaceGDB, isPercentFlood)
                                    #print(surfaceRasterProcessList)
                                    if error == 1:
                                        arcpy.AddMessage("Could not detect risk values in naming of surface rasters. Continuing processing without Hydro Surface Attributes")
                                        surfaceDataExists = False
                                    else:
                                        if len(surfaceRasterProcessList) > 0:
                                            surfaceDataExists = True
                                            inWaterSurfaceType = "Multipatch"
                                        else:
                                            surfaceDataExists = False
                                else:
                                    # check if rasters are present
                                    riskSurfaceRasters = flood_impact_lib.obtain_rasters(inSurfaceGDB)
                                    if len(riskSurfaceRasters) > 0:  # Detect that Raster Data Exists

                                        # alert user if data will not be processed: risk value needs to be present in naming.
                                        surfaceRasterProcessList, error = flood_impact_lib.obtain_processing_list(riskSurfaceRasters, riskValues, inSurfaceGDB, isPercentFlood)
                                        #print(surfaceRasterProcessList)
                                        if error == 1:
                                            arcpy.AddMessage("Could not detect risk values in naming of surface rasters. Continuing processing without Hydro Surface Attributes")
                                            surfaceDataExists = False
                                        else:
                                            if len(surfaceRasterProcessList) > 0:
                                                surfaceDataExists = True
                                                inWaterSurfaceType = "Raster"
                                            else:
                                                surfaceDataExists = False
                                    else:
                                        arcpy.AddMessage("No Surface Rasters detected. Continuing processing without Hydro Surface Attributes")
                                        surfaceDataExists = False
                            else:
                                arcpy.AddMessage("No Surface Rasters detected. Continuing processing without Hydro Surface Attributes")
                                surfaceDataExists = False

                            # check if depth data exists...
                            if arcpy.Exists(inDepthGDB):
                                # Obtain List of Available Depth Features/Rasters in GDB
                                riskDepthRasters = flood_impact_lib.obtain_rasters(inDepthGDB)

                                depthRasterProcessList, error = flood_impact_lib.obtain_processing_list(riskDepthRasters, riskValues, inDepthGDB, isPercentFlood)

                                if error == 1:
                                    arcpy.AddWarning("Could not detect risk values in naming of depth rasters. Check your raster names and chosen riskType Exiting...")
                                    inDepthGDB = ""
                                    raise InputError
                                else:
                                    # Check that all depth rasters are positive values.
                                    for raster in depthRasterProcessList:
                                        rasterMinDepth = arcpy.GetRasterProperties_management(raster[2],
                                                                                              "MINIMUM").getOutput(0)

                                        rasterMinDepth = float(re.sub("[,.]", ".", rasterMinDepth))

                                        if rasterMinDepth < 0:
                                            arcpy.AddError(
                                                "Depth Rasters must not contain negative values. Terminating process")
                                            arcpy.AddWarning(
                                                "Please check rasters in your depth raster geodatabase to ensure they are all positive depth values")
                                            raise InputError

                                    # Ensure consistencies with the Depth And Surface Raster Datasets
                                    ''' Process will automatically exit processing and report issue if inconsistent data'''
                                    if surfaceDataExists:
                                        compareDepthSurfaceLists(surfaceRasterProcessList, depthRasterProcessList)
                            else:
                                arcpy.AddWarning("Detected that Depth Rasters were non-existant.\n"
                                                 "Process will generate Depth Rasters using Surface Raster and DEM.\n"
                                                 "Note: The resulting Attributes will not be accurate.. only relative")
                                inDepthGDB = ""

                            if len(inDepthGDB) > 0 and len(depthRasterProcessList) > 0:
                                arcpy.SetProgressor("default", "Obtaining vector Geometry Information for Processing")
                                exposureFieldsList = []

                                if lc_use_in_memory:
                                    tempFP = os.path.join("in_memory", "tempFP")
                                else:
                                    tempFP = os.path.join(scratch_ws, "tempFP")

                                createTempFP(inFeature, bufferDistance, featureFID, tempFP)
                                # Convert Footprint to raster
                                arcpy.SetProgressor("default", "Commencing Rasterization of Vector Features for Processing")

                                if lc_use_in_memory:
                                    tempRasterFP = os.path.join("in_memory", "tempRasterFP")
                                else:
                                    # zonal stats fails on in memory raster
                                    tempRasterFP = os.path.join(scratch_ws, "tempRasterFP")

                                if arcpy.Exists(tempRasterFP):
                                    arcpy.Delete_management(tempRasterFP)

                                # Set Pixel Size for All Raster Analysis as the Smallest Input Raster Size. Ensures Accuracy
                                arcpy.PolygonToRaster_conversion(tempFP, featureFID, tempRasterFP, "CELL_CENTER", "NONE",
                                                                 tolerance)
                                arcpy.env.cellSize = "MINOF"

                                # Define Pixel Area for Footprint
                                PIXELAREA = pixelArea(tempRasterFP)

                                # Begin Calculating Depth Statistics Information
                                # Check for Meters vs Feet elevations in DEM vs Water Surface Rasters
                                count = 0
                                exposureLevels = len(depthRasterProcessList)

                                # for each depth raster in gdb, process and add stats to outTable
                                if 1: #REMOVE!!!!!!!!!!!!
                                    for row in depthRasterProcessList:
                                        arcpy.SetProgressor("step", "Step progressor: Processing Exposure Level {0} of {1}".format(
                                            count, exposureLevels), count, exposureLevels, 1)
                                        arcpy.SetProgressorPosition(count)
                                        arcpy.AddMessage("Round {0} of {1}".format(count+1, exposureLevels))

                                        if count >= 0:  # For testing on only the First File Set; Remove upon completion
                                            surfaceRaster = None

                                            if inWaterSurfaceType == "Multipatch":
                                                # Look-Up the Feature of interest in surfaceRasterProcessList
                                                riskGeom3D = surfaceRasterProcessList[count][2]  # Row 2 = the multipatch location

                                                if lc_use_in_memory:
                                                    surfaceRaster = os.path.join("in_memory",
                                                                                 "SurfaceRaster{0}{1}".format(riskValues[0],
                                                                                                              row[0]))
                                                else:
                                                    surfaceRaster = os.path.join(scratch_ws,
                                                                                 "SurfaceRaster{0}{1}".format(riskValues[0],
                                                                                                              row[0]))
                                                if arcpy.Exists(surfaceRaster):
                                                    arcpy.Delete_management(surfaceRaster)

                                                MpRasterTool(riskGeom3D, surfaceRaster)

                                            if inWaterSurfaceType == "Raster":
                                                surfaceRaster = surfaceRasterProcessList[count][2]  # Row 2 = the Raster Location

                                            # get raster name

                                            depthRaster = depthRasterProcessList[count][2]

                                            depthZonalStatsTable = os.path.join(scratch_ws, "depthZonalStatsTable{0}{1}".format(riskValues[0], row[0]))

                                            # Begin Calculating Depth Statistics Information
                                            if arcpy.Exists(depthZonalStatsTable):
                                                arcpy.Delete_management(depthZonalStatsTable)

                                            arcpy.AddMessage("Calculating Depth Statistics Information for " + common_lib.get_name_from_feature_class(depthRaster) + ".")
                                            arcpy.sa.ZonalStatisticsAsTable(tempRasterFP, featureFID, depthRaster, depthZonalStatsTable,
                                                                            "DATA", "ALL")
                                            fields = arcpy.ListFields(depthZonalStatsTable)
                                            deleteFieldList = ['ZONE_CODE']
                                            for field in fields:
                                                if str(field.name) in deleteFieldList:
                                                    arcpy.DeleteField_management(depthZonalStatsTable, field.name)
                                                else:
                                                    if str(field.name) not in ['OID', 'OBJECTID', featureFID, "Shape_Area"]:
                                                        arcpy.AlterField_management(depthZonalStatsTable, field.name,
                                                                                    '{0}{1}{2}'.format(riskValues[0], row[0], field.name),
                                                                                    '{0} {1} {2}'.format(riskValues[0], row[0], field.name))

                                            ''' Add Geometry Area attribute to the highest Surface Table for use in future calcs...'''
#                                            arcpy.AddField_management(depthZonalStatsTable, "Shape_Area", "DOUBLE", None, None,
#                                                                      None,
#                                                                      "Shape Area", "NULLABLE", "NON_REQUIRED", None)

                                            ''' simply join the shape_area attribute created by the buffer --> no need to calculate it below '''
                                            arcpy.JoinField_management(depthZonalStatsTable, featureFID, tempFP, featureFID, "Shape_Area")

#                                            with arcpy.da.UpdateCursor(depthZonalStatsTable,
#                                                                       [featureFID, "Shape_Area"]) as ucursor:
#                                                for uRow in ucursor:
#                                                    with arcpy.da.SearchCursor(tempFP, [featureFID, "SHAPE@AREA"]) as scursor:
#                                                        for sRow in scursor:
#                                                            if uRow[0] == sRow[0]:
#                                                                # sRow[0].area  # Read area value as double
#                                                                uRow[1] = sRow[1]  # Write area value to field
#                                                                ucursor.updateRow(uRow)


                                            # Overwrite Area field with Count*(PixelLength*PixelWidth)
                                            strAreaField = '{0}{1}{2}'.format(riskValues[0], row[0], 'AREA')
                                            countField = '{0}{1}{2}'.format(riskValues[0], row[0], 'COUNT')
                                            sumField = '{0}{1}{2}'.format(riskValues[0], row[0], 'SUM')

                                            with arcpy.da.UpdateCursor(depthZonalStatsTable, [strAreaField, countField, "Shape_Area"]) as ucursor:
                                                for item in ucursor:
                                                    exposedArea = item[1] * PIXELAREA

                                                    if exposedArea > item[2]:
                                                        item[0] = item[2]
                                                    else:
                                                        item[0] = exposedArea
                                                    ucursor.updateRow(item)

                                            # Delete Count Field as no longer necessary
                                            if 1: #DeleteIntermediateData:
                                                arcpy.DeleteField_management(depthZonalStatsTable, countField)

                                            # Add Exposure Attribute to table and calculate Percent Exposure
                                            strExposureField = "{0}{1}Exposure".format(riskValues[0], row[0])
                                            exposureFieldsList.append(strExposureField)

                                            arcpy.AddMessage("Adding fields...")
                                            arcpy.AddField_management(depthZonalStatsTable, strExposureField, "DOUBLE", None,
                                                                      None, None, "{0} {1} Exposure".format(riskValues[0], row[0]),
                                                                      "NULLABLE",
                                                                      "NON_REQUIRED", None)

                                            arcpy.AddMessage("Updating fields...")
                                            with arcpy.da.UpdateCursor(depthZonalStatsTable,
                                                                       [strExposureField, strAreaField, "Shape_Area"]) as ucursor:
                                                for uRow in ucursor:
                                                    val = round((100/uRow[2]) * uRow[1], 4)  # Write percent exposed area value to field
                                                    if val > 100:  # Ensure percent exposure caps at 100 percent.
                                                        val = 100
                                                    uRow[0] = val
                                                    ucursor.updateRow(uRow)

                                            # Calculate Water Volume
                                            if volumeField:
                                                arcpy.AddMessage("Updating volume field...")

                                                volumeField = "{0}{1}Volume".format(riskValues[0], row[0])
                                                arcpy.AddField_management(depthZonalStatsTable, volumeField, "DOUBLE", None,
                                                                          None, None, "{0} {1} Volume".format(riskValues[0], row[0]),
                                                                          "NULLABLE",
                                                                          "NON_REQUIRED", None)
                                                with arcpy.da.UpdateCursor(depthZonalStatsTable,
                                                                           [volumeField, strAreaField, sumField]) as ucursor:
                                                    for uRow in ucursor:
                                                        # Define Volume Equation
                                                        volume = round(uRow[2] * PIXELAREA, 4)

                                                        uRow[0] = volume
                                                        ucursor.updateRow(uRow)

                                            # calculate loss potential
                                            if lossTable:
                                                if lossField:
                                                    arcpy.AddMessage("Updating loss potential field...")

                                                    lossField = "{0}{1}LossPotential".format(riskValues[0], row[0])
                                                    averageField = '{0}{1}{2}'.format(riskValues[0], row[0], 'MEAN')

                                                    # we assume same number of values for each field
                                                    depth_list = sorted(common_lib.get_row_values_for_fields(None, loss_gdb_table, [depthField], None, "no_expression"))
                                                    loss_list = sorted(common_lib.get_row_values_for_fields(None, loss_gdb_table, [potentialLossField], None, "no_expression"))
                                                    size_list = sorted(common_lib.get_row_values_for_fields(None, loss_gdb_table, [sizeField], None, "no_expression"))

                                                    if len(depth_list) == len(loss_list) == len(size_list):
                                                        arcpy.AddField_management(depthZonalStatsTable, lossField, "DOUBLE", None,
                                                                                   None, None, "{0} {1} LossPotential".format(riskValues[0], row[0]),
                                                                                   "NULLABLE",
                                                                                   "NON_REQUIRED", None)

                                                        with arcpy.da.UpdateCursor(depthZonalStatsTable,
                                                                                    [lossField, averageField, strAreaField, "Shape_Area"]) as ucursor:
                                                            i = 0
                                                            for uRow in ucursor:
                                                                #    get loss value for mean depth
                                                                closest_depth, index = common_lib.find_closest(depth_list, uRow[1])

                                                                # use same index to find loss and size values
                                                                potential_loss = loss_list[index]
                                                                size = size_list[index]

                                                                # calculate loss based on shape_area: loss from table * flooded area / area from table
                                                                # if area is zero, we take the full amount.
                                                                if size == 0:
                                                                    loss = potential_loss
                                                                else:
                                                                    loss = (potential_loss * uRow[2]) / size

                                                                uRow[0] = loss
                                                                ucursor.updateRow(uRow)
                                                                i+=1

                                                                pass
                                                    else:
                                                        print("Error reading LossPotential Table. Missing values. Exiting...")
                                                        arcpy.AddError("Error reading LossPotential Table. Missing values. Exiting...")

                                            # Delete Sum Field as no longer necessary
                                            arcpy.DeleteField_management(depthZonalStatsTable, sumField)

                                            # Delete
                                            delFieldList = [
                                                [areaField, '{0}{1}{2}'.format(riskValues[0], row[0], 'AREA')],
                                                [minField, '{0}{1}{2}'.format(riskValues[0], row[0], 'MIN')],
                                                [maxField, '{0}{1}{2}'.format(riskValues[0], row[0], 'MAX')],
                                                [rangeField, '{0}{1}{2}'.format(riskValues[0], row[0], 'RANGE')],
                                                [meanField, '{0}{1}{2}'.format(riskValues[0], row[0], 'MEAN')],
                                                [stdField, '{0}{1}{2}'.format(riskValues[0], row[0], 'STD')]
                                            ]
                                            for field in delFieldList:
                                                if field[0] is False:
                                                    for fieldName in arcpy.ListFields(depthZonalStatsTable):
                                                        # print(str(field[1]), str(fieldName.name))
                                                        if str(field[1]) == str(fieldName.name):
                                                            arcpy.DeleteField_management(depthZonalStatsTable, fieldName.name)

                                            # Check for Existance of WSEL Raster GDB:
                                            if WaterSurfaceElevationLevel:
                                                if arcpy.Exists(inSurfaceGDB):
                                                    arcpy.AddMessage("Updating water surface elevation field...")

                                                    # Parse Mean Water Surface Elevation Values to feature:
                                                    surfaceZonalStatsTable = os.path.join("in_memory",
                                                                                        "surfaceZonalStatsTable{0}{1}".format(riskValues[0], row[0]))

                                                    # Begin Calculating Surface Statistics Information as WSEL
                                                    # Generates new table to be appended upon processing
                                                    if arcpy.Exists(surfaceZonalStatsTable):
                                                        arcpy.Delete_management(surfaceZonalStatsTable)
                                                    arcpy.sa.ZonalStatisticsAsTable(tempRasterFP, featureFID, surfaceRaster,
                                                                                    surfaceZonalStatsTable, "DATA", "MEAN")
                                                    fields = arcpy.ListFields(surfaceZonalStatsTable)
                                                    deleteFieldList = ['ZONE_CODE', 'COUNT', 'AREA']
                                                    for field in fields:
                                                        if str(field.name) in deleteFieldList:
                                                            arcpy.DeleteField_management(surfaceZonalStatsTable, field.name)
                                                        else:
                                                            if str(field.name) not in ['OID', 'OBJECTID', featureFID, "Shape_Area"]:
                                                                arcpy.AlterField_management(
                                                                    surfaceZonalStatsTable,
                                                                    field.name,
                                                                    '{0}{1}{2}'.format(riskValues[0], row[0], "WSEL"),
                                                                    '{0} {1} {2}'.format(riskValues[0], row[0], "WSEL")
                                                                )
                                                    arcpy.JoinField_management(depthZonalStatsTable, featureFID, surfaceZonalStatsTable,
                                                                               featureFID, '{0}{1}{2}'.format(riskValues[0], row[0],
                                                                                                              "WSEL"))
                                                    arcpy.Delete_management(surfaceZonalStatsTable)

                                            # Delete Intermediate Raster Data From Generated Derivatives where certain data did not exist
                                            if inWaterSurfaceType == "Multipatch":
                                                try:
                                                    arcpy.Delete_management(surfaceRaster)
                                                except:
                                                    arcpy.AddWarning("Could not delete {0}".format(surfaceRaster))
                                                    pass
                                            if not arcpy.Exists(inDepthGDB):
                                                try:
                                                    arcpy.Delete_management(depthRaster)
                                                except:
                                                    arcpy.AddWarning("Could not delete {0}".format(depthRaster))

                                            #  Copy the First Table as a "Master Table" for merging all additional tables to...
                                            if count == 0:
                                                # All additional tables will be appended to this table
                                                if arcpy.Exists(outTable):
                                                    arcpy.Delete_management(outTable)
                                                outGDB = os.path.split(outTable)[0]
                                                outTableName = os.path.split(outTable)[1]

                                                arcpy.AddMessage("Creating risk table...")

                                                arcpy.TableToTable_conversion(depthZonalStatsTable, outGDB, outTableName)
                                                arcpy.Delete_management(depthZonalStatsTable)
                                            else:
                                                field_names = [f.name for f in arcpy.ListFields(depthZonalStatsTable)]
                                                removeFields = ["OID", "OBJECTID", "Shape_Area", "Shape Area", featureFID]
                                                for field in removeFields:
                                                    if field in field_names:
                                                        field_names.remove(field)
                                                arcpy.JoinField_management(outTable, featureFID, depthZonalStatsTable, featureFID,
                                                                           field_names)

                                                arcpy.AddMessage("Appending to risk table...")

                                                arcpy.Delete_management(depthZonalStatsTable)
                                        count += 1

                                    arcpy.SetProgressor("default", "Calculating Exposure Level of First Impact")
                                    if not shapeAreaField:
                                        arcpy.DeleteField_management(outTable, 'Shape_Area')

                                    # Calculate the Risk Slider Value each feature is first exposed at ... Begins at 0:
                                    arcpy.AddField_management(outTable, '{0}{1}'.format(riskValues[0], "Slider"), "SHORT", None, None, None,
                                                              "{0} {1}".format(riskValues[0], "Slider"), "NULLABLE", "NON_REQUIRED", None)
                                    inList = []
                                    inList.extend(exposureFieldsList)
                                    inList.append('{0}{1}'.format(riskValues[0], "Slider"))
                                    with arcpy.da.UpdateCursor(outTable, inList) as ucursor:
                                        for uRow in ucursor:
                                            itemCount = 0
                                            firstExposedLevel = None
                                            for exposure in exposureFieldsList:
                                                if uRow[itemCount] is not None:
                                                    firstExposedLevel = (len(exposureFieldsList) - 1) - itemCount
#                                                    firstExposedLevel = len(exposureFieldsList) - itemCount
                                                itemCount += 1
                                            uRow[len(uRow) - 1] = firstExposedLevel
                                            ucursor.updateRow(uRow)
                                            del itemCount
                                    del inList
                                    del exposureFieldsList

                                    # Calculate the true Risk Level each feature is first exposed at:
                                    arcpy.AddField_management(outTable, '{0}{1}'.format(riskValues[0], "Level"), "DOUBLE", None, None, None,
                                                              "{0} {1}".format(riskValues[0], "Level"), "NULLABLE", "NON_REQUIRED", None)
                                    inList = ['{0}{1}'.format(riskValues[0], "Slider"),
                                              '{0}{1}'.format(riskValues[0], "Level")]
                                    with arcpy.da.UpdateCursor(outTable, inList) as ucursor:
                                        for uRow in ucursor:
                                            if uRow[0] is not None:
                                                uRow[1] = riskValues[1][::-1][int(uRow[0])]
                                                ucursor.updateRow(uRow)
                                    del inList

                                # check if there is any output

                                num_features = arcpy.GetCount_management(outTable).getOutput(0)

                                if int(num_features) > 0:
                                    # output a summary table for use in Operations Dashboard - temporary until there is a custom app
                                    # Create summary stats table
                                    temp_solution = 1
                                    if temp_solution:
                                        sum_stats_fields_plus_type_list = []
                                        sum_stats_fields_list = []

                                        # Loop through all fields in the outTable Table and define the ones to run stats on
                                        for field in arcpy.ListFields(outTable):
                                            # Just find the fields that have a numeric type
                                            if field.type in ("Double", "Integer", "Single", "SmallInteger", "Float"):
                                                # Add the field name and Sum statistic type
                                                #    to the list of fields to summarize
                                                if riskValues[0] in field.name and "AREA" in field.name:
                                                    sum_stats_fields_plus_type_list.append([field.name, "Sum"])
                                                    sum_stats_fields_list.append("SUM_" + field.name)
                                                if riskValues[0] in field.name and "LossPotential" in field.name:
                                                    sum_stats_fields_plus_type_list.append([field.name, "Sum"])
                                                    sum_stats_fields_list.append("SUM_" + field.name)
                                                if riskValues[0] in field.name and "Volume" in field.name:
                                                    sum_stats_fields_plus_type_list.append([field.name, "Sum"])
                                                    sum_stats_fields_list.append("SUM_" + field.name)

                                        if lc_use_in_memory:
                                            outTableStats = "in_memory/outTableStats"
                                        else:
                                            outTableStats = os.path.join(scratch_ws, "outTableStats")
                                            if arcpy.Exists(outTableStats):
                                                arcpy.Delete_management(outTableStats)

                                        # sum stats on area, loss_potential and volume
                                        arcpy.Statistics_analysis(outTable, outTableStats, sum_stats_fields_plus_type_list)

                                        # create table with indicator per level
                                        # create as many features as there are risk levels to serve as outStatsTable
                                        ex_string, xmin, xmax, ymin, ymax = obtainExtent(inFeature)

                                        points_list = []
                                        p = 0
                                        for p in range(exposureLevels + 1):     # 1 extra to account for zero normal situation
                                            points_list.append(((xmin + p, ymin + p),))
                                            p += 1

                                        sr = arcpy.Describe(inFeature).spatialReference
                                        stats_table = outTable + "_stats"

                                        if arcpy.Exists(stats_table):
                                            arcpy.Delete_management(stats_table)

                                        # create actual point stats feature class
                                        create_point_featureclass_from_points(stats_table, points_list, sr)

                                        # calculate buildings affected per level and write to table
                                        # use select by attribute on level: level <= range levels
                                        number_affected_field = "number_affected"
                                        damage_area_field = "damage_area"
                                        loss_potential_field = "loss_potential"
                                        volume_field = "volume"
                                        slider_field = "slider"
                                        level_field = "level"
                                        stats_field_dict = {}
                                        stats_field_dict.update({number_affected_field: "LONG"})
                                        stats_field_dict.update({damage_area_field: "FLOAT"})
                                        stats_field_dict.update({loss_potential_field: "FLOAT"})
                                        stats_field_dict.update({volume_field: "FLOAT"})
                                        stats_field_dict.update({volume_field: "FLOAT"})
                                        stats_field_dict.update({slider_field: "LONG"})

                                        arcpy.AddMessage("Adding required fields to statistics table...")

                                        stats_field_list = []
                                        for k, v in stats_field_dict.items():
                                            common_lib.delete_add_field(stats_table, k, v)
                                            stats_field_list.append(k)

                                        # get list of number affected per level
                                        number_affected_list = []
                                        number_affected_list.append(0)      # 0 == normal no flood

                                        for ba in range(exposureLevels):
                                            slider_field = '{0}{1}'.format(riskValues[0], "Slider")
                                            expression = """{} <= {}""".format(arcpy.AddFieldDelimiters(outTable, slider_field), ba)
                                            number = common_lib.get_feature_count(outTable, expression)
                                            number_affected_list.append(number)

                                        # get list of damage per level
                                        damage_list = []
                                        loss_potential_list = []
                                        volume_list = []
                                        have_loss_potential = False

                                        # we assume descending list of attributes and there should be exposureLevel number of fields for all 3 sum_stats fields
                                        for field in sum_stats_fields_list:
                                            if "AREA" in field:
                                                damage_area_value = common_lib.get_row_values_for_fields(None, outTableStats,
                                                                                                      [field],
                                                                                                      None, "no_expression")
                                                damage_list.append(damage_area_value)

                                            if "LossPotential" in field:
                                                loss_potential_value = common_lib.get_row_values_for_fields(None,
                                                                                                            outTableStats,
                                                                                                            [field],
                                                                                                            None,
                                                                                                            "no_expression")
                                                loss_potential_list.append(loss_potential_value)
                                                have_loss_potential = True

                                            if "Volume" in field:
                                                volume_value = common_lib.get_row_values_for_fields(None, outTableStats,
                                                                                                   [field],
                                                                                                   None, "no_expression")
                                                volume_list.append(volume_value)

                                        # reverse lists to ascending for processing
                                        damage_list.append([0])  # 0 == normal no flood
                                        damage_list = damage_list[::-1]
                                        loss_potential_list.append([0])  # 0 == normal no flood
                                        loss_potential_list = loss_potential_list[::-1]
                                        volume_list.append([0])  # 0 == normal no flood
                                        volume_list = volume_list[::-1]

                                        # update colums in stats_table
                                        i = 0
                                        with arcpy.da.UpdateCursor(stats_table, stats_field_list) as ucursor:
                                            for uRow in ucursor:
                                                uRow[0] = number_affected_list[i]
                                                uRow[1] = damage_list[i][0]
                                                if have_loss_potential:
                                                    uRow[2] = loss_potential_list[i][0]
                                                else:
                                                    uRow[2] = 0
                                                uRow[3] = volume_list[i][0]
                                                uRow[4] = i

                                                ucursor.updateRow(uRow)
                                                i += 1

                                    # end temp_solution

                                    # Remove Exposure Field
                                    if exposureField is False:
                                        for fieldName in arcpy.ListFields(outTable):
                                            if 'Exposure' in str(fieldName.name):
                                                arcpy.DeleteField_management(outTable, fieldName.name)
                                            if 'SUM' in str(fieldName.name):
                                                arcpy.DeleteField_management(outTable, fieldName.name)

                                    # Check for Existance of DEM Raster
                                    # Calculate DEM Elevation Statistics on Features
                                    if arcpy.Exists(inDEM):
                                        if isRaster(inDEM):
                                            arcpy.SetProgressor("default", "Attributing Ground Elevation Statistics")
                                            demZonalStatsTable = os.path.join("in_memory", "demZonalStatsTable")
                                            # Begin Calculating Depth Statistics Information
                                            if arcpy.Exists(demZonalStatsTable):
                                                arcpy.Delete_management(demZonalStatsTable)
                                            arcpy.sa.ZonalStatisticsAsTable(tempRasterFP, featureFID, inDEM,
                                                                            demZonalStatsTable,
                                                                            "DATA", "ALL")
                                            fields = arcpy.ListFields(demZonalStatsTable)
                                            deleteFieldList = ['ZONE_CODE', 'SUM', 'AREA', 'COUNT']
                                            delFieldList = [
                                                [GroundMinField, 'MIN'],
                                                [GroundMaxField, 'MAX'],
                                                [GroundRangeField, 'RANGE'],
                                                [GroundMeanField, 'MEAN'],
                                                [GroundSTDField, 'STD']
                                            ]
                                            for field in delFieldList:
                                                if field[0] is False:
                                                    for fieldName in arcpy.ListFields(demZonalStatsTable):
                                                        if str(field[1]) == str(fieldName.name):
                                                            arcpy.DeleteField_management(demZonalStatsTable, fieldName.name)
                                            fields = arcpy.ListFields(demZonalStatsTable)
                                            for field in fields:
                                                if str(field.name) in deleteFieldList:
                                                    arcpy.DeleteField_management(demZonalStatsTable, field.name)
                                                else:
                                                    if str(field.name) not in ['OID', 'OBJECTID', featureFID, "Shape_Area", 'SUM']:
                                                        arcpy.AlterField_management(demZonalStatsTable,
                                                                                    field.name,
                                                                                    '{0}{1}'.format("DEM", field.name),
                                                                                    '{0} {1}'.format("DEM", field.name))
                                                        arcpy.JoinField_management(outTable,
                                                                                   featureFID,
                                                                                   demZonalStatsTable,
                                                                                   featureFID,
                                                                                   '{0}{1}'.format("DEM", field.name))
                                            arcpy.Delete_management(demZonalStatsTable)

                                    # we have succeeded!
                                    success = True

                                    # Detect Time Required to Complete Process
                                    end_time = time.clock()
                                    msg_body = create_msg_body("attribute_feature completed successfully.", start_time,
                                                               end_time)
                                    msg(msg_body)

                                else:
                                    # Detect Time Required to Complete Process
                                    end_time = time.clock()
                                    msg_body = create_msg_body("attribute_feature failed.", start_time,
                                                               end_time)
                                    msg(msg_body)

                                    success = False
                                    stats_table = None

                                if DeleteIntermediateData:
                                    # Delete Temporary Footprint
                                    arcpy.Delete_management(tempFP)
                                    # arcpy.Delete_management(tempRasterFP)

                                # Delete Data Stored "In_Memory"
                                arcpy.Delete_management("in_memory")

                                # Clear the Workspace Cache....
                                arcpy.ClearWorkspaceCache_management()

                                return success, inFeature, stats_table

                            else:
                                arcpy.AddError("Can't find any Depth Rasters. Exiting...")
                                arcpy.AddWarning("Please check naming of rasters in your depth raster geodatabase. See documentation for correct naming protocol.")
                                raise InputError
                        else:
                            msg_body = create_msg_body("Can't read: " + inSpreadsheet + "!", 0, 0)
                            msg(msg_body, ERROR)

                        # Clear the Workspace Cache....
                        arcpy.ClearWorkspaceCache_management()

                    else:
                        raise NoRiskTable
                else:
                    raise GeopraphicSR
            else:
                raise LicenseErrorSpatial
        else:
            raise LicenseError3D
        arcpy.ClearWorkspaceCache_management()

    except NotProjected:
        print("Input data needs to be in a projected coordinate system. Exiting...")
        arcpy.AddError("Input data needs to be in a projected coordinate system. Exiting...")

    except NoLayerFile:
        print("Can't find Layer file. Exiting...")
        arcpy.AddError("Can't find Layer file. Exiting...")

    except LicenseError3D:
        print("3D Analyst license is unavailable")
        arcpy.AddError("3D Analyst license is unavailable")

    except LicenseErrorSpatial:
        print("Spatial Analyst license is unavailable")
        arcpy.AddError("Spatial Analyst license is unavailable")

    except NoNoDataError:
        print("Input raster does not have NODATA values")
        arcpy.AddError("Input raster does not have NODATA values")

    except NoUnits:
        print("No units detected on input data")
        arcpy.AddError("No units detected on input data")

    except GeopraphicSR:
        # The input has geographic SR
        #
        print((
                  'Input data has geographic coordinate system. Only projected coordinate systems are supported. Also ensure all input is in the same projected spatial reference, including the same vertical units.'))
        arcpy.AddError(
            'Input data has geographic coordinate system. Only projected coordinate systems are supported. Also ensure all input is in the same projected spatial reference, including the same vertical units.')

    except MixOfSR:
        # The input has mixed SR
        #
        print((
                  'Input data has mixed spatial references. Ensure all input is in the same projected spatial reference, including the same vertical units.'))
        arcpy.AddError(
            'Input data has mixed spatial references. Ensure all input is in the same projected spatial reference, including the same vertical units.')

    except NoPolygons:
        print("Input data can only be polygon features or raster datasets.")
        arcpy.AddError("Input data can only be polygon features or raster datasets.")

    except NoRiskTable:
        print("Can't find RiskTable: riskTypeTable.xlsx in the tables directory.")
        arcpy.AddError("Can't find RiskTable: riskTypeTable.xlsx in the tables directory.")

    except NoLossTable:
        print("Error with LossPotential Table. Exiting...")
        arcpy.AddError("Error with LossPotential Table. Exiting...")

    except ValueError:
        print("Input no flood value is not a number.")
        arcpy.AddError("Input no flood value is not a number.")

    except InputError:
        print("Inconsistent input.")
        arcpy.AddError("Inconsistent input.")
        return False, None, None

    except arcpy.ExecuteError:
        line, filename, synerror = trace()
        msg("Error on %s" % line, ERROR)
        msg("Error in file name:  %s" % filename, ERROR)
        msg("With error message:  %s" % synerror, ERROR)
        msg("ArcPy Error Message:  %s" % arcpy.GetMessages(2), ERROR)

    except FunctionError as f_e:
        messages = f_e.args[0]
        msg("Error in function:  %s" % messages["function"], ERROR)
        msg("Error on %s" % messages["line"], ERROR)
        msg("Error in file name:  %s" % messages["filename"], ERROR)
        msg("With error message:  %s" % messages["synerror"], ERROR)
        msg("ArcPy Error Message:  %s" % messages["arc"], ERROR)

    except:
        line, filename, synerror = trace()
        msg("Error on %s" % line, ERROR)
        msg("Error in file name:  %s" % filename, ERROR)
        msg("with error message:  %s" % synerror, ERROR)

#    finally:
#        arcpy.CheckInExtension("3D")
#        arcpy.CheckInExtension("Spatial")


# for debug only!
if __name__ == "__main__":
    riskType = "Storm Surge"  # "NOAA Sea Level Rise", "FEMA Flood", "Tidal Flood", "Storm Surge", "Riverine Flood"
    isPercentFlood = True
    inWaterSurfaceType = ""  # "Raster", "Multipatch"
    inSurfaceGDB = r'D:\Gert\Work\Esri\Solutions\3DFloodImpact\data\test\test_WSE.gdb'
    inDepthGDB = r'D:\Temp\Flood\3DFloodImpact\test_depth.gdb'
    inFeature = r'D:\Gert\Work\Esri\Solutions\3DFloodImpact\work2.2\FloodImpactPlanning_old\TestData\Test.gdb\test_buildings_mp'
    featureFID = "OBJECTID"
    bufferDistance = 2
    tolerance = 1
    inDEM = r''
    outGDB = r'D:\Gert\Work\Esri\Solutions\3DFloodImpact\work2.2\FloodImpactPlanning_old\Testing.gdb'
    outTable = r'D:\Gert\Work\Esri\Solutions\3DFloodImpact\work2.2\FloodImpactPlanning_old\Testing.gdb\riskTable_Buildings'
    inLossTable = r'D:\Temp\Flood\3DFloodImpact\tables\fema_loss_potential_meter.xls\fema_loss_potential$'

    areaField = True
    minField = True
    maxField = True
    rangeField = True
    meanField = True
    stdField = True
    volumeField = True  # Stable
    shapeAreaField = True
    exposureField = True
    WaterSurfaceElevationLevel = True
    GroundMinField = True
    GroundMaxField = True
    GroundRangeField = True
    GroundMeanField = True
    GroundSTDField = True
    lossField = True

    esri_featureID = "copy_featureID"

    # create string field for featureFID. otherwise resulting raster won't be integer
    common_lib.delete_add_field(inFeature, esri_featureID, "TEXT")
    arcpy.CalculateField_management(inFeature, esri_featureID, "!" + featureFID + "!", "PYTHON_9.3")

    # make the featureFID the text version of featureFID
    featureFID = esri_featureID

    attribute_feature(riskType=riskType,
                      isPercentFlood=isPercentFlood,
                      inWaterSurfaceType=inWaterSurfaceType,
                      inSurfaceGDB=inSurfaceGDB,
                      inDepthGDB=inDepthGDB,
                      inFeature=inFeature,
                      featureFID=featureFID,
                      bufferDistance=bufferDistance,
                      tolerance=tolerance,
                      inDEM=inDEM,
                      lossTable=inLossTable,
                      outTable=outTable,
                      areaField=areaField,
                      minField=minField,
                      maxField=maxField,
                      rangeField=rangeField,
                      meanField=meanField,
                      stdField=stdField,
                      volumeField=volumeField,
                      shapeAreaField=shapeAreaField,
                      exposureField=exposureField,
                      WaterSurfaceElevationLevel=WaterSurfaceElevationLevel,
                      GroundMinField=GroundMinField,
                      GroundMaxField=GroundMaxField,
                      GroundRangeField=GroundRangeField,
                      GroundMeanField=GroundMeanField,
                      GroundSTDField=GroundSTDField,
                      lossField=lossField,
                      debug=1,
                      lc_use_in_memory=False)
